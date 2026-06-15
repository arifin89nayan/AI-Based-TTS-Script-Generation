"""
The Fixed List backed by an Excel (.xlsx) file.

Columns (matches the source Excel format): Serial No | Word | Phoneme

Word    = surface (kanji word)
Phoneme = reading (katakana pronunciation)

New entries are appended at the bottom with the next Serial number.
The original file layout and existing rows are never overwritten.
"""
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

_HEADERS = ["Serial No", "Word", "Phoneme"]


class FixedList:
    def __init__(self, path: str):
        self._path = Path(path)
        # Word -> {"Serial No", "Word", "Phoneme"}
        self._data: Dict[str, dict] = {}
        # ordered list of Word keys so Serial numbering stays stable
        self._order: List[str] = []
        self._load()

    # ------------------------------------------------------------------
    # internal helpers
    # ------------------------------------------------------------------

    def _load(self) -> None:
        if not self._path.exists():
            self._save()
            return
        wb = openpyxl.load_workbook(self._path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            wb.close()
            return
        headers = [str(h).strip() if h else "" for h in rows[0]]
        for row in rows[1:]:
            entry = dict(zip(headers, row))
            word = str(entry.get("Word", "") or "").strip()
            if word:
                self._data[word] = {
                    "Serial No": entry.get("Serial No", ""),
                    "Word": word,
                    "Phoneme": str(entry.get("Phoneme", "") or "").strip(),
                }
                if word not in self._order:
                    self._order.append(word)
        wb.close()

    def _save(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "FixedList"

        # header row — bold, centred
        ws.append(_HEADERS)
        header_font = Font(bold=True)
        center = Alignment(horizontal="center")
        for col in range(1, len(_HEADERS) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.alignment = center

        # data rows in insertion order
        for i, word in enumerate(self._order, start=1):
            entry = self._data[word]
            ws.append([i, entry["Word"], entry["Phoneme"]])
            ws.cell(row=i + 1, column=1).alignment = center  # Serial No centred

        # column widths
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 36

        wb.save(self._path)
        wb.close()

    # ------------------------------------------------------------------
    # public interface
    # ------------------------------------------------------------------

    def contains(self, surface: str) -> bool:
        return surface in self._data

    def get(self, surface: str) -> Optional[str]:
        entry = self._data.get(surface)
        return entry["Phoneme"] if entry else None

    def add(self, surface: str, reading: str) -> None:
        """Append a new entry (or update existing), then write to disk."""
        if surface not in self._data:
            self._order.append(surface)
        self._data[surface] = {
            "Serial No": len(self._order),
            "Word": surface,
            "Phoneme": reading,
        }
        self._save()

    def all(self) -> Dict[str, str]:
        return {word: entry["Phoneme"] for word, entry in self._data.items()}

    def close(self) -> None:
        pass    # data is persisted on every add(); nothing to flush here
